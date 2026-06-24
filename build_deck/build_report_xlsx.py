"""Генерира Excel model report по формата от курса (notes.md)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PLUM = "7B2D5E"
PINK = "F2D7E3"
WHITE = "FFFFFF"
FONT = "Arial"

wb = Workbook()
ws = wb.active
ws.title = "Model report"

thin = Side(style="thin", color="C9A9BC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style(cell, *, bold=False, size=11, color="000000", fill=None, align="left", wrap=False):
    cell.font = Font(name=FONT, bold=bold, size=size, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fill:
        cell.fill = PatternFill("solid", start_color=fill)


# --- заглавие ---
ws.merge_cells("A1:J1")
style(ws["A1"], bold=True, size=14, color=PLUM, align="left")
ws["A1"] = "Model report — Разпознаване на медицинска аномалия (CAMELYON17)"

# --- кой е най-добрият модел и защо ---
ws.merge_cells("A3:J3")
style(ws["A3"], bold=True, size=11, color="000000", fill=PINK, align="left", wrap=True)
ws["A3"] = ("Най-добър модел: VGG11 — най-висок Cohen Kappa (0.797) върху невиждани "
            "пациенти, което по Landis & Koch е значително съгласие. По-дълбокият VGG16 "
            "не дава предимство при този обем данни.")
ws.row_dimensions[3].height = 32

# --- хедър на таблицата ---
headers = ["Модел", "Архитектура", "Скорост\nна обуч.", "Епохи", "Аугментация",
           "Accuracy\n(val)", "Δ% спрямо\nbaseline", "Cohen Kappa\n(val)",
           "Δ спрямо\nbaseline", "Comments"]
for j, h in enumerate(headers, 1):
    c = ws.cell(row=5, column=j, value=h)
    style(c, bold=True, color=WHITE, fill=PLUM, align="center", wrap=True)
    c.border = border
ws.row_dimensions[5].height = 30

# --- редове: всеки ред е модел (хипотеза), в реда на създаване ---
# (acc, kappa са измерени; baseline = мажоритарен класификатор на балансиран набор)
rows = [
    ["Baseline (мнозинствен клас)", "—", "—", "—", "—", 0.50, None, 0.00, None,
     "Предсказва само мнозинствения клас. Референтна точка."],
    ["SimpleCNN", "3 conv блока", "5e-4", 25, "ColorJitter + флипове", 0.863, None, 0.725, None,
     "Базова мрежа; солидно над baseline."],
    ["VGG11", "VGG11 + BN", "5e-4", 25, "ColorJitter + флипове", 0.899, None, 0.797, None,
     "Най-добър. Значително съгласие (Landis & Koch)."],
    ["VGG16", "VGG16 + BN", "5e-4", 25, "ColorJitter + флипове", 0.862, None, 0.724, None,
     "По-дълбок, без предимство при този обем данни."],
]
for i, r in enumerate(rows):
    row = 6 + i
    for j, val in enumerate(r, 1):
        c = ws.cell(row=row, column=j, value=val)
        c.border = border
        style(c, align="center" if j not in (1, 10) else "left", wrap=(j == 10))
    # формули за промяната спрямо baseline (baseline е ред 6)
    if row > 6:
        g = ws.cell(row=row, column=7, value=f"=(F{row}-$F$6)/$F$6")
        g.number_format = "+0.0%;-0.0%"
        g.border = border
        style(g, align="center")
        k = ws.cell(row=row, column=9, value=f"=H{row}-$H$6")
        k.number_format = "+0.000;-0.000"
        k.border = border
        style(k, align="center")
    else:
        ws.cell(row=row, column=7, value="baseline").border = border
        ws.cell(row=row, column=9, value="baseline").border = border
        style(ws.cell(row=row, column=7), align="center")
        style(ws.cell(row=row, column=9), align="center")
    ws.cell(row=row, column=6).number_format = "0.000"
    ws.cell(row=row, column=8).number_format = "0.000"

# подчертай най-добрия (VGG11 = ред 8)
for j in range(1, 11):
    c = ws.cell(row=8, column=j)
    c.fill = PatternFill("solid", start_color=PINK)
    c.font = Font(name=FONT, bold=True, size=11)

# --- бележки под таблицата ---
ws.merge_cells("A11:J11")
style(ws["A11"], size=10, color="555555", wrap=True)
ws["A11"] = ("U-Net автоенкодер (unsupervised): ROC AUC на грешката на реконструкция = "
             "0.353 (< 0.5 — не разделя класовете; туморните патчове се реконструират "
             "по-добре). Не е включен в таблицата, защото метриката му не е сравнима.")
ws.row_dimensions[11].height = 28
ws.merge_cells("A12:J12")
style(ws["A12"], size=10, color="555555", wrap=True)
ws["A12"] = ("Оценката е върху валидационно множество от 12 000 патча от пациенти извън "
             "обучението (разделяне на ниво пациент). Отделно тестово множество е бъдеща "
             "работа. Метрика на Cohen е реализирана от нула в metrics.py.")
ws.row_dimensions[12].height = 28

widths = [26, 14, 9, 7, 18, 10, 11, 12, 11, 42]
for j, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64 + j)].width = w

# --- лист 2: confusion matrix на най-добрия модел (VGG11) ---
cm = wb.create_sheet("Confusion matrix VGG11")
style(cm["A1"], bold=True, size=12, color=PLUM)
cm["A1"] = "VGG11 — матрица на обърквания (валидация, 12 000 патча)"
labels = ["", "Предсказан normal", "Предсказан tumor", "Общо", "Дял верни"]
for j, h in enumerate(labels, 1):
    c = cm.cell(row=3, column=j, value=h)
    style(c, bold=True, color=WHITE, fill=PLUM, align="center", wrap=True)
    c.border = border
data = [["Истински normal", 5842, 158], ["Истински tumor", 1058, 4942]]
for i, r in enumerate(data):
    row = 4 + i
    for j, v in enumerate(r, 1):
        c = cm.cell(row=row, column=j, value=v)
        c.border = border
        style(c, bold=(j == 1), align="left" if j == 1 else "center")
    cm.cell(row=row, column=4, value=f"=B{row}+C{row}").border = border  # общо
    rate = cm.cell(row=row, column=5, value=f"={'B' if i == 0 else 'C'}{row}/D{row}")
    rate.number_format = "0.0%"
    rate.border = border
    style(cm.cell(row=row, column=4), align="center")
    style(rate, align="center")
style(cm["A7"], size=10, color="555555", wrap=True)
cm.merge_cells("A7:E7")
cm["A7"] = ("Специфичност 97.4% (рядко обявява здрава тъкан за тумор); чувствителност "
            "82.4% (пропуска 17.6% от тумора). Грешките са асиметрични — фалшиво "
            "отрицателните (пропуснат тумор) са по-съществени в медицински контекст.")
cm.row_dimensions[7].height = 40
for j, w in enumerate([18, 18, 18, 10, 12], 1):
    cm.column_dimensions[chr(64 + j)].width = w

wb.save("../docs/model_report.xlsx")
print("записан docs/model_report.xlsx")
